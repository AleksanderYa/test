from openpyxl import load_workbook, Workbook
  
wb = load_workbook(filename = 'Pricess.xlsx')
sheet_ranges = wb['cc']


wb2 = load_workbook(filename = 'Pricess2.xlsx')
sheet_ranges2 = wb2['cc']

price_dict = {}


def write(ryad , col, value):
    try:
        wb = openpyxl.load_workbook(filename='C:\\Users\\Sales2\\Desktop\\test\\Pricess2.xlsx' )
        ws = wb.worksheets[0]
        ws.cell(ryad , col).value = value
        wb.save('C:\\Users\\Sales2\\Desktop\\test\\Pricess2.xlsx')
        return print(ws.cell(ryad , col).value,'ok')
    except Exception as e:
        return None
        print(e)
        
for i in range(1, 50):
    value = sheet_ranges['P{0}'.format(i)].value
    if value and type(value) == int:
        price_dict[value] = sheet_ranges['L{0}'.format(i)].value
        print(value, sheet_ranges['L{0}'.format(i)].value)

print(price_dict.keys())
    
for i in range(8, 28):
    value = sheet_ranges2['S{0}'.format(i)].value
    if value in price_dict.keys():
        try:
            write(12, i, price_dict[i])
            #print(price_dict[i])
        except Exception as e:
            print(e)
        
 





""" no
wb3 = Workbook()

dest_filename = 'Prices.xlsx'

ws1 = wb3.active
ws1.title = "cc"

for row in range(1, 40):
    ws1.append(range(600))

#ws2 = wb.create_sheet(title="Pi")

wb3.save(filename = dest_filename)
"""



""" no
def find_price():   
    for i in range(2, 21):
        yield [sheet_ranges['P{0}'.format(i)].value, i]
    
print('In Pricess2')

for i in range(8, 28):
    # print(sheet_ranges2['S{0}'.format(i)].value)
    find_price = find_price()
    try:
        for ii in range(25):
            find = next(find_price)
            if sheet_ranges2['S{0}'.format(i)].value == find[0]:
                print(i, find[1])
    except Exception as e:
        print(e)
"""

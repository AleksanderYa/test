from xlrd import open_workbook
import datetime
import openpyxl
import xlwt

a = str(datetime.datetime.today())
date = a[8:10] + ',' + a[5:7] + ',' + a[2:4]
print(date)
book = open_workbook('C:\\Users\Sales2\Desktop\Александр&БАЗА\ВсеКлиентыДляБазыsales2\Остальное\КурсВалют.xls')
sheet = book.sheet_by_index(0)

def dolars():
    from xlrd import open_workbook    
    book = open_workbook('C:\\Users\Sales2\Desktop\Александр&БАЗА\ВсеКлиентыДляБазыsales2\Остальное\КурсВалют.xls')
    sheet = book.sheet_by_index(0)
    try:    
        for i in range(10000):
            if sheet.row_values(i)[5]:
                if sheet.row_values(i)[5] == date:
                    dolars = sheet.row_values(i)[6]
                    print(dolars)
            else:
                print('dich')
                break
    except Exception as e:
        print(e)
    return dolars
dolars = dolars()
 
def dolars_ots():
    from xlrd import open_workbook    
    book = open_workbook('C:\\Users\Sales2\Desktop\Александр&БАЗА\ВсеКлиентыДляБазыsales2\Остальное\КурсВалют.xls')
    sheet = book.sheet_by_index(0)
    try:    
        for i in range(10000):
            if sheet.row_values(i)[5]:
                if sheet.row_values(i)[5] == date:
                    dolars_ots = sheet.row_values(i)[7]
            else:
                break
    except Exception as e:
        print(e)
    return dolars_ots
dolars_ots = dolars_ots()




def evro():
    from xlrd import open_workbook    
    book = open_workbook('C:\\Users\Sales2\Desktop\Александр&БАЗА\ВсеКлиентыДляБазыsales2\Остальное\КурсВалют.xls')
    sheet = book.sheet_by_index(0)
    try:    
        for i in range(1000):
            if sheet.row_values(i)[5]:
                if sheet.row_values(i)[5] == date:
                    evro = sheet.row_values(i)[8]
            else:
                break
    except Exception as e:
#         print(e)
        pass
    return evro
evro = evro()
 
def evro_ots():
    from xlrd import open_workbook    
    book = open_workbook('C:\\Users\Sales2\Desktop\Александр&БАЗА\ВсеКлиентыДляБазыsales2\Остальное\КурсВалют.xls')
    sheet = book.sheet_by_index(0)
    try:    
        for i in range(1000):
            if sheet.row_values(i)[5]:
                if sheet.row_values(i)[5] == date:
                    evro_ots = sheet.row_values(i)[9]
            else:
                break
    except:
        pass
    return evro_ots
evro_ots = evro_ots()




book2 = open_workbook('C:\\Users\Sales2\Desktop\Pricess.xlsx')
sheet2 = book2.sheet_by_index(0)

# books = open_workbook('C:\\Users\Sales2\Desktop\PricessSublime.xlsx')
# sheets = books.sheet_by_index(0)
try:
    def write(ryad , col, value):  
        wb = openpyxl.load_workbook(filename='C:\\Users\\Sales2\Desktop\Pricess.xlsx' )
        ws = wb.worksheets[0]
        ws.cell(ryad , col).value = value
        wb.save('C:\\Users\\Sales2\Desktop\Pricess.xlsx')
        return print(ws.cell(ryad , col).value,'ok')
except Exception as e:
    print(e)
    
# def writesublime(ryad , col, value):  
    # wbs = openpyxl.load_workbook(filename='C:\\Users\\Sales2\Desktop\PricessSublime.xlsx')
    # wss = wbs.worksheets[0]
    # wss.cell(ryad , col).value = value
    # wbs.save('C:\\Users\\Sales2\Desktop\PricessSublime.xlsx')
    # return print(wss.cell(ryad , col).value,'ok')


try:
    write(21, 4, date)
    write(22, 6, dolars)
    write(22, 7, dolars_ots)
    write(23, 6, evro)
    write(23, 7, evro_ots)
    for i in range(10000):
        if sheet2.row_values(i)[10]:
#             print(sheet2.row_values(i)[10])
            if sheet2.row_values(i)[10] == 'evro':
#                 print(evro)
                evros = sheet2.row_values(i)[9] * evro
                evros_opt = sheet2.row_values(i)[8] * evro
#                 print(evros)
                evros = '{:.2f}'.format(evros)
                evros_opt = '{:.2f}'.format(evros_opt)                
                evros_ots = sheet2.row_values(i)[9] * evro_ots
                evros_ots_opt = sheet2.row_values(i)[8] * evro_ots
                evros_ots= '{:.2f}'.format(evros_ots) 
                evros_ots_opt= '{:.2f}'.format(evros_ots_opt)
                write(i+1, 12, evros)
                write(i+1, 13, evros_ots)
                write(i+1, 14, evros_opt)
                write(i+1, 15, evros_ots_opt)
#                 evros_ots = sheet.row_values(i)[9] * evro_ots
#                 print(evros, evros_ots)
                
            elif sheet2.row_values(i)[10] == 'dolar':
                dolarse = sheet2.row_values(i)[9] * dolars
                dolarse_opt = sheet2.row_values(i)[8] * dolars
                dolarse = '{:.2f}'.format(dolarse)
                dolarse_opt = '{:.2f}'.format(dolarse_opt)
                dolarse_ots = sheet2.row_values(i)[9] * dolars_ots
                dolarse_ots_opt = sheet2.row_values(i)[8] * dolars_ots
                dolarse_ots = '{:.2f}'.format(dolarse_ots)
                dolarse_ots_opt = '{:.2f}'.format(dolarse_ots_opt)
#                 print(dolarse, dolarse_ots)
                write(i+1, 12, dolarse)
                write(i+1, 13, dolarse_ots)
                write(i+1, 14, dolarse_opt)
                write(i+1, 15, dolarse_ots_opt)
    # for i in range(10000):
        # if sheets.row_values(i)[10]:
            # print(sheet2.row_values(i)[10])
            # if sheets.row_values(i)[10] == 'evro':
                # print(evro)
                # evross = sheets.row_values(i)[11] * evro
                # print(evros)
                # evross = '{:.2f}'.format(evros)  
                # evros_ots_sub = sheets.row_values(i)[11] * evro_ots
                # evros_ots_sub = '{:.2f}'.format(evros_ots_sub)  
                # writesublime(i+1, 12, evross)
                # writesublime(i+1, 13, evros_ots_sub)
                # evros_ots = sheet.row_values(i)[11] * evro_ots
                # print(evros, evros_ots)
            
        
except Exception as e:
    print(e,'?')
    

    




# def write_in_file(col, ryad, value):
#     # Работает ! Єто запись в файл    
#     wb = openpyxl.load_workbook(filename='C:\\Users\Sales2\Desktop\Prices.xlsx' )
#     ws = wb.worksheets[0]
#     ws.cell(ryad , col).value = value
# #     while True:
# #         o = next(one)
# #         if not ws.cell(o , 1).value:
# #             nomb = int(ws.cell(int(o)-1 , 1).value)
# #             print(nomb)
# #             ws.cell(o , 1).value =  nomb + 1
# #             break

#     wb.save('C:\\Users\Sales2\Desktop\Prices.xlsx')
